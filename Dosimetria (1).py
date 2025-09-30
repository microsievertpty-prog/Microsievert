# -*- coding: utf-8 -*-
import re
import io
import unicodedata
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import requests
import streamlit as st
from datetime import datetime

# ====== Excel helpers (openpyxl) ======
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ===================== NINOX CONFIG =====================
API_TOKEN   = "edf312a0-98b8-11f0-883e-db77626d62e5"
TEAM_ID     = "YrsYfTegptdZcHJEj"
DATABASE_ID = "ow1geqnkz00e"
BASE_URL    = "https://api.ninox.com/v1"
TABLE_WRITE_NAME = "BASE DE DATOS"

# ===================== UI =====================
st.set_page_config(page_title="Microsievert — Dosimetría", page_icon="🧪", layout="wide")
st.title("🧪 Carga y Cruce de Dosis → Ninox (**BASE DE DATOS**)")

# ===================== Helpers =====================
def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def hp_to_num(x) -> float:
    if x is None:
        return 0.0
    s = str(x).strip().upper()
    if s in ("", "PM", "NONE", "NAN"):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0

def pmfmt2(v, thr: float = 0.005) -> str:
    try:
        f = float(v)
    except Exception:
        s = str(v).strip()
        return "PM" if s == "" else s
    return "PM" if f < thr else f"{f:.2f}"

def last_nonempty(series: pd.Series) -> str:
    for v in series.iloc[::-1]:
        if str(v).strip():
            return str(v)
    return ""

def is_control_name(x: str) -> bool:
    s = strip_accents(str(x or "")).upper()
    s = re.sub(r"\s+", " ", s).strip()
    return s.startswith("CONTROL")

def safe_cols(df: pd.DataFrame, cols: List[str]) -> List[str]:
    return [c for c in cols if c in df.columns]

# ---------- Normalización de PERIODO ----------
MES_MAP = {
    "ENE":"ENERO","FEB":"FEBRERO","MAR":"MARZO","ABR":"ABRIL","MAY":"MAYO","JUN":"JUNIO",
    "JUL":"JULIO","AGO":"AGOSTO","SEP":"SEPTIEMBRE","OCT":"OCTUBRE","NOV":"NOVIEMBRE","DIC":"DICIEMBRE",
    "JAN":"ENERO","APR":"ABRIL","AUG":"AGOSTO","DEC":"DICIEMBRE",
}
MES_NUM = {"01":"ENERO","02":"FEBRERO","03":"MARZO","04":"ABRIL","05":"MAYO","06":"JUNIO",
           "07":"JULIO","08":"AGOSTO","09":"SEPTIEMBRE","10":"OCTUBRE","11":"NOVIEMBRE","12":"DICIEMBRE"}

def _to_year4(y: str) -> str:
    y = y.strip()
    return f"20{y}" if len(y) == 2 else y

def normalizar_periodo(valor: str) -> str:
    if not valor:
        return ""
    s = strip_accents(str(valor)).upper().strip()
    s = re.sub(r"\s+", " ", s)

    m = re.match(r"^(JAN|ENE|FEB|MAR|APR|ABR|MAY|JUN|JUL|AUG|AGO|SEP|OCT|NOV|DEC|DIC)[\s\-/]*([0-9]{2,4})$", s)
    if m:
        mes = MES_MAP.get(m.group(1), m.group(1))
        anio = _to_year4(m.group(2))
        return f"{mes} {anio}"

    m = re.match(r"^([0-1][0-9])[\s\-/]*([0-9]{2,4})$", s)
    if m and m.group(1) in MES_NUM:
        mes = MES_NUM[m.group(1)]
        anio = _to_year4(m.group(2))
        return f"{mes} {anio}"

    m = re.match(r"^([0-9]{4})[\s\-/]*([0-1][0-9])$", s)
    if m and m.group(2) in MES_NUM:
        mes = MES_NUM[m.group(2)]
        anio = m.group(1)
        return f"{mes} {anio}"

    return s

MES_A_NUM = {"ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,"JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12}
def periodo_to_date(s: str):
    if not s or not isinstance(s, str):
        return pd.NaT
    s = s.strip().upper()
    m = re.match(r"^(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)\s+([0-9]{4})$", s)
    if not m:
        return pd.NaT
    mes = MES_A_NUM.get(m.group(1))
    an = int(m.group(2))
    try:
        return pd.Timestamp(year=an, month=mes, day=1)
    except Exception:
        return pd.NaT

# ===================== Ninox helpers =====================
def ninox_headers():
    return {"Authorization": f"Bearer {API_TOKEN}", "Content-Type": "application/json"}

@st.cache_data(ttl=300, show_spinner=False)
def ninox_list_tables(team_id: str, db_id: str):
    url = f"{BASE_URL}/teams/{team_id}/databases/{db_id}/tables"
    r = requests.get(url, headers=ninox_headers(), timeout=30)
    r.raise_for_status()
    return r.json()

def resolve_table_id(table_hint: str) -> str:
    hint = (table_hint or "").strip()
    if hint and " " not in hint and len(hint) <= 8:
        return hint
    for t in ninox_list_tables(TEAM_ID, DATABASE_ID):
        if str(t.get("name", "")).strip().lower() == hint.lower():
            return str(t.get("id", "")).strip()
    return hint

def ninox_insert(table_hint: str, rows: List[Dict[str, Any]], batch_size: int = 300) -> Dict[str, Any]:
    table_id = resolve_table_id(table_hint)
    url = f"{BASE_URL}/teams/{TEAM_ID}/databases/{DATABASE_ID}/tables/{table_id}/records"
    n, inserted = len(rows), 0
    if n == 0:
        return {"ok": True, "inserted": 0}
    for i in range(0, n, batch_size):
        chunk = rows[i:i+batch_size]
        r = requests.post(url, headers=ninox_headers(), json=chunk, timeout=60)
        if r.status_code != 200:
            return {"ok": False, "inserted": inserted, "error": f"{r.status_code} {r.text}"}
        inserted += len(chunk)
    return {"ok": True, "inserted": inserted}

@st.cache_data(ttl=300, show_spinner=False)
def ninox_list_records(table_hint: str, limit: int = 1000, max_pages: int = 50):
    table_id = resolve_table_id(table_hint)
    url = f"{BASE_URL}/teams/{TEAM_ID}/databases/{DATABASE_ID}/tables/{table_id}/records"
    out: List[Dict[str, Any]] = []
    skip = 0
    for _ in range(max_pages):
        params = {"limit": limit, "skip": skip}
        r = requests.get(url, headers=ninox_headers(), params=params, timeout=60)
        r.raise_for_status()
        batch = r.json() or []
        if not batch:
            break
        out.extend(batch)
        if len(batch) < limit:
            break
        skip += limit
    return out

def ninox_records_to_df(records: List[Dict[str,Any]]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame()
    rows = []
    for rec in records:
        f = rec.get("fields", {}) or {}
        rows.append({
            "PERIODO DE LECTURA": f.get("PERIODO DE LECTURA"),
            "CLIENTE": f.get("CLIENTE"),
            "CÓDIGO DE DOSÍMETRO": f.get("CÓDIGO DE DOSÍMETRO"),
            "CÓDIGO DE USUARIO": f.get("CÓDIGO DE USUARIO"),
            "NOMBRE": f.get("NOMBRE"),
            "CÉDULA": f.get("CÉDULA"),
            "FECHA DE LECTURA": f.get("FECHA DE LECTURA"),
            "TIPO DE DOSÍMETRO": f.get("TIPO DE DOSÍMETRO"),
            "Hp (10)": f.get("Hp (10)"),
            "Hp (0.07)": f.get("Hp (0.07)"),
            "Hp (3)": f.get("Hp (3)"),
        })
    df = pd.DataFrame(rows)
    if "PERIODO DE LECTURA" in df.columns:
        df["PERIODO DE LECTURA"] = df["PERIODO DE LECTURA"].astype(str).map(normalizar_periodo)
    return df

# ===================== Lectores de archivos =====================
def norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    def _n(x: str) -> str:
        x = strip_accents(str(x)).strip()
        x = re.sub(r"\s+", " ", x)
        return x
    out = df.copy()
    out.columns = [_n(c) for c in out.columns]
    return out

def coalesce_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    got = [c for c in df.columns if c in candidates]
    return got[0] if got else None

def parse_csv_robust(upload) -> pd.DataFrame:
    raw = upload.read(); upload.seek(0)
    for sep in [";", ",", None]:
        for enc in ["utf-8-sig", "latin-1"]:
            try:
                return pd.read_csv(BytesIO(raw), sep=sep, engine="python", encoding=enc)
            except Exception:
                continue
    try:
        return pd.read_excel(BytesIO(raw))
    except Exception:
        raise

def leer_lista_codigo(upload) -> Optional[pd.DataFrame]:
    if not upload: return None
    name = upload.name.lower()
    if name.endswith((".xlsx", ".xls")):
        xls = pd.ExcelFile(upload)
        sheet = None
        for s in xls.sheet_names:
            s_norm = strip_accents(s).lower()
            if "asignar" in s_norm and "dosimet" in s_norm:
                sheet = s; break
        if sheet is None:
            sheet = xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sheet)
    else:
        df = parse_csv_robust(upload)

    df = norm_cols(df)
    c_ced   = coalesce_col(df, ["CEDULA"])
    c_user  = coalesce_col(df, ["CODIGO DE USUARIO","CODIGO_USUARIO","CODIGO USUARIO"])
    c_nom   = coalesce_col(df, ["NOMBRE","NOMBRE COMPLETO","NOMBRE_COMPLETO"])
    c_ap    = coalesce_col(df, ["APELLIDO","APELLIDOS"])
    c_cli   = coalesce_col(df, ["CLIENTE","COMPANIA"])
    c_cod   = coalesce_col(df, ["CODIGO_DOSIMETRO","CODIGO DE DOSIMETRO","CODIGO DOSIMETRO"])
    c_per   = coalesce_col(df, ["PERIODO DE LECTURA","PERIODO_DE_LECTURA","PERIODO"])
    c_tipo  = coalesce_col(df, ["TIPO DE DOSIMETRO","TIPO_DE_DOSIMETRO","TIPO DOSIMETRO"])
    c_etq   = coalesce_col(df, ["ETIQUETA"])

    out = pd.DataFrame()
    out["CÉDULA"]            = df[c_ced] if c_ced else ""
    out["CÓDIGO DE USUARIO"] = df[c_user] if c_user else ""
    if c_nom and c_ap:
        out["NOMBRE"] = (df[c_nom].astype(str).str.strip() + " " + df[c_ap].astype(str).str.strip()).str.strip()
    elif c_nom:
        out["NOMBRE"] = df[c_nom].astype(str).str.strip()
    else:
        out["NOMBRE"] = ""
    out["CLIENTE"]           = df[c_cli].astype(str).str.strip() if c_cli else ""
    out["CÓDIGO_DOSÍMETRO"]  = (df[c_cod].astype(str).str.strip().str.upper() if c_cod else "")
    out["PERIODO DE LECTURA"] = (df[c_per].astype(str).map(normalizar_periodo) if c_per else "")
    out["TIPO DE DOSÍMETRO"] = df[c_tipo].astype(str).str.strip() if c_tipo else ""
    out["ETIQUETA"]          = df[c_etq].astype(str).str.strip() if c_etq else ""

    def _is_ctrl(r):
        return is_control_name(r.get("NOMBRE","")) or is_control_name(r.get("ETIQUETA",""))

    out["_IS_CONTROL"] = out.apply(_is_ctrl, axis=1)
    return out

def leer_dosis(upload) -> Optional[pd.DataFrame]:
    if not upload: return None
    name = upload.name.lower()
    if name.endswith((".xlsx",".xls")):
        df = pd.read_excel(upload)
    else:
        df = parse_csv_robust(upload)

    cols = (df.columns.astype(str).str.strip().str.lower()
            .str.replace(" ", "", regex=False)
            .str.replace("(", "", regex=False).str.replace(")", "", regex=False)
            .str.replace(".", "", regex=False))
    df.columns = cols

    if "dosimeter" not in df.columns:
        for alt in ["dosimetro","codigo","codigodosimetro","codigo_dosimetro"]:
            if alt in df.columns:
                df.rename(columns={alt:"dosimeter"}, inplace=True); break

    for cands, dest in [ (["hp10dosecorr","hp10dose","hp10"], "hp10dose"),
                         (["hp007dosecorr","hp007dose","hp007"], "hp0.07dose"),
                         (["hp3dosecorr","hp3dose","hp3"], "hp3dose") ]:
        for c in cands:
            if c in df.columns: df.rename(columns={c:dest}, inplace=True); break
        if dest not in df.columns: df[dest] = 0.0
        else: df[dest] = pd.to_numeric(df[dest], errors="coerce").fillna(0.0)

    if "dosimeter" in df.columns:
        df["dosimeter"] = df["dosimeter"].astype(str).str.strip().str.upper()
    if "timestamp" in df.columns:
        df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
    return df

# ===================== Construcción de registros =====================
def construir_registros(df_lista: pd.DataFrame,
                        df_dosis: pd.DataFrame,
                        periodos: List[str]) -> pd.DataFrame:
    df_l = df_lista.copy()
    df_l["PERIODO DE LECTURA"] = df_l["PERIODO DE LECTURA"].astype(str).str.strip().str.upper()
    df_l["CÓDIGO_DOSÍMETRO"]   = df_l["CÓDIGO_DOSÍMETRO"].astype(str).str.strip().str.upper()

    selected = [p.strip().upper() for p in periodos if str(p).strip()]
    if selected:
        df_l = df_l[df_l["PERIODO DE LECTURA"].isin(selected)]

    idx = df_dosis.set_index("dosimeter") if "dosimeter" in df_dosis.columns else pd.DataFrame().set_index(pd.Index([]))

    registros: List[Dict[str, Any]] = []
    df_l = pd.concat([df_l[df_l["_IS_CONTROL"]], df_l[~df_l["_IS_CONTROL"]]], ignore_index=True)

    for _, r in df_l.iterrows():
        cod = str(r["CÓDIGO_DOSÍMETRO"]).strip().upper()
        if not cod or cod == "NAN":
            continue
        if cod not in idx.index:
            continue
        d = idx.loc[cod]
        if isinstance(d, pd.DataFrame):
            d = d.sort_values(by="timestamp").iloc[-1]
        ts = d.get("timestamp", pd.NaT)
        fecha_str = ""
        try:
            fecha_str = pd.to_datetime(ts).strftime("%d/%m/%Y %H:%M") if pd.notna(ts) else ""
        except Exception:
            fecha_str = ""

        nombre = str(r.get("NOMBRE","")).strip()
        if bool(r["_IS_CONTROL"]) and (not nombre):
            nombre = "CONTROL"

        registros.append({
            "PERIODO DE LECTURA": r["PERIODO DE LECTURA"],
            "CLIENTE": r.get("CLIENTE",""),
            "CÓDIGO DE DOSÍMETRO": cod,
            "CÓDIGO DE USUARIO": r.get("CÓDIGO DE USUARIO",""),
            "NOMBRE": nombre,
            "CÉDULA": r.get("CÉDULA",""),
            "FECHA DE LECTURA": fecha_str,
            "TIPO DE DOSÍMETRO": r.get("TIPO DE DOSÍMETRO","") or "CE",
            "Hp (10)":  float(d.get("hp10dose", 0.0) or 0.0),
            "Hp (0.07)":float(d.get("hp0.07dose", 0.0) or 0.0),
            "Hp (3)":   float(d.get("hp3dose", 0.0) or 0.0),
            "_IS_CONTROL": bool(r["_IS_CONTROL"])
        })

    df_final = pd.DataFrame(registros)
    if not df_final.empty:
        df_final = df_final.sort_values(["_IS_CONTROL","NOMBRE","CÉDULA"], ascending=[False, True, True]).reset_index(drop=True)
    return df_final

# ===================== Resta de CONTROL + Formato =====================
def aplicar_resta_control_y_formato(
    df_final: pd.DataFrame,
    umbral_pm: float = 0.005,
    manual_ctrl: Optional[float] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if df_final is None or df_final.empty:
        return df_final, df_final

    df = df_final.copy()
    for h in ["Hp (10)", "Hp (0.07)", "Hp (3)"]:
        if h not in df.columns:
            df[h] = 0.0
        df[h] = pd.to_numeric(df[h], errors="coerce").fillna(0.0)
    if "PERIODO DE LECTURA" in df.columns:
        df["PERIODO DE LECTURA"] = df["PERIODO DE LECTURA"].astype(str).map(normalizar_periodo)

    is_ctrl = df["NOMBRE"].apply(is_control_name)
    df_ctrl = df[is_ctrl].copy()
    df_per  = df[~is_ctrl].copy()

    ctrl_means = pd.DataFrame(columns=["PERIODO DE LECTURA","Hp10_CTRL","Hp007_CTRL","Hp3_CTRL"])
    if not df_ctrl.empty:
        ctrl_means = df_ctrl.groupby("PERIODO DE LECTURA", as_index=False).agg({
            "Hp (10)":"mean","Hp (0.07)":"mean","Hp (3)":"mean"
        }).rename(columns={"Hp (10)":"Hp10_CTRL","Hp (0.07)":"Hp007_CTRL","Hp (3)":"Hp3_CTRL"})

    out = df_per.copy()
    if not ctrl_means.empty:
        out = out.merge(ctrl_means, on="PERIODO DE LECTURA", how="left")
        out["Hp10_CTRL"]  = out["Hp10_CTRL"].fillna(0.0)
        out["Hp007_CTRL"] = out["Hp007_CTRL"].fillna(0.0)
        out["Hp3_CTRL"]   = out["Hp3_CTRL"].fillna(0.0)
        out["_Hp10_NUM"]  = (out["Hp (10)"]   - out["Hp10_CTRL"]).clip(lower=0.0)
        out["_Hp007_NUM"] = (out["Hp (0.07)"] - out["Hp007_CTRL"]).clip(lower=0.0)
        out["_Hp3_NUM"]   = (out["Hp (3)"]    - out["Hp3_CTRL"]).clip(lower=0.0)
    else:
        if manual_ctrl is not None and float(manual_ctrl) > 0:
            cval = float(manual_ctrl)
            out["_Hp10_NUM"]  = (out["Hp (10)"]   - cval).clip(lower=0.0)
            out["_Hp007_NUM"] = (out["Hp (0.07)"] - cval).clip(lower=0.0)
            out["_Hp3_NUM"]   = (out["Hp (3)"]    - cval).clip(lower=0.0)
        else:
            out["_Hp10_NUM"]  = out["Hp (10)"]
            out["_Hp007_NUM"] = out["Hp (0.07)"]
            out["_Hp3_NUM"]   = out["Hp (3)"]

    out_view = out.copy()
    out_view["Hp (10)"]   = out_view["_Hp10_NUM"].map(lambda v: pmfmt2(v, umbral_pm))
    out_view["Hp (0.07)"] = out_view["_Hp007_NUM"].map(lambda v: pmfmt2(v, umbral_pm))
    out_view["Hp (3)"]    = out_view["_Hp3_NUM"].map(lambda v: pmfmt2(v, umbral_pm))

    df_ctrl_view = pd.DataFrame()
    if not df_ctrl.empty:
        df_ctrl_view = df_ctrl.merge(ctrl_means, on="PERIODO DE LECTURA", how="left")
        for c in ["Hp10_CTRL","Hp007_CTRL","Hp3_CTRL"]:
            df_ctrl_view[c] = df_ctrl_view[c].fillna(0.0)
        df_ctrl_view["_Hp10_NUM"]  = (df_ctrl_view["Hp (10)"]   - df_ctrl_view["Hp10_CTRL"]).clip(lower=0.0)
        df_ctrl_view["_Hp007_NUM"] = (df_ctrl_view["Hp (0.07)"] - df_ctrl_view["Hp007_CTRL"]).clip(lower=0.0)
        df_ctrl_view["_Hp3_NUM"]   = (df_ctrl_view["Hp (3)"]    - df_ctrl_view["Hp3_CTRL"]).clip(lower=0.0)
        df_ctrl_view["Hp (10)"]    = df_ctrl_view["_Hp10_NUM"].map(lambda v: pmfmt2(v, umbral_pm))
        df_ctrl_view["Hp (0.07)"]  = df_ctrl_view["_Hp007_NUM"].map(lambda v: pmfmt2(v, umbral_pm))
        df_ctrl_view["Hp (3)"]     = df_ctrl_view["_Hp3_NUM"].map(lambda v: pmfmt2(v, umbral_pm))

    df_vista = pd.concat([df_ctrl_view, out_view], ignore_index=True, sort=False)
    if not df_vista.empty:
        df_vista["__is_control__"] = df_vista["NOMBRE"].apply(is_control_name)
        df_vista = df_vista.sort_values(by=["__is_control__","NOMBRE","CÉDULA"], ascending=[False, True, True]).drop(columns=["__is_control__"])

    df_num = out[[ "_Hp10_NUM","_Hp007_NUM","_Hp3_NUM","PERIODO DE LECTURA","CLIENTE",
                   "CÓDIGO DE USUARIO","CÓDIGO DE DOSÍMETRO","NOMBRE","CÉDULA",
                   "TIPO DE DOSÍMETRO","FECHA DE LECTURA" ]].copy()

    return df_vista, df_num

# ===================== Consolidación para subir =====================
def consolidar_para_upload(df_vista: pd.DataFrame, df_num: pd.DataFrame, umbral_pm: float = 0.005) -> pd.DataFrame:
    if df_vista is None or df_vista.empty or df_num is None or df_num.empty:
        return pd.DataFrame()

    # PERSONAS
    personas_num = df_num[~df_num["NOMBRE"].apply(is_control_name)].copy()
    per_consol = pd.DataFrame()
    if not personas_num.empty:
        per_consol = personas_num.groupby(["PERIODO DE LECTURA","CÓDIGO DE USUARIO"], as_index=False).agg({
            "CLIENTE":"last","NOMBRE":"last","CÉDULA":"last","CÓDIGO DE DOSÍMETRO":"last",
            "TIPO DE DOSÍMETRO":"last","FECHA DE LECTURA":"last",
            "_Hp10_NUM":"sum","_Hp007_NUM":"sum","_Hp3_NUM":"sum"
        }).rename(columns={"_Hp10_NUM":"Hp (10)","_Hp007_NUM":"Hp (0.07)","_Hp3_NUM":"Hp (3)"})

    # CONTROL
    control_v = df_vista[df_vista["NOMBRE"].apply(is_control_name)].copy()
    ctrl_consol = pd.DataFrame()
    if not control_v.empty:
        for h in ["Hp (10)","Hp (0.07)","Hp (3)"]:
            control_v[h] = control_v[h].apply(hp_to_num)

        def _last_nonempty(series: pd.Series) -> str:
            for v in series.iloc[::-1]:
                s = str(v).strip()
                if s:
                    return s
            return ""

        ctrl_consol = control_v.groupby(["PERIODO DE LECTURA"], as_index=False).agg({
            "CLIENTE":"last",
            "CÓDIGO DE DOSÍMETRO":"first",
            "CÓDIGO DE USUARIO": _last_nonempty,
            "CÉDULA": _last_nonempty,
            "TIPO DE DOSÍMETRO":"last",
            "FECHA DE LECTURA":"last",
            "Hp (10)":"mean","Hp (0.07)":"mean","Hp (3)":"mean"
        })

        def _fill_usercode(row):
            cu = str(row.get("CÓDIGO DE USUARIO","") or "").strip()
            if cu:
                return cu
            cd = str(row.get("CÓDIGO DE DOSÍMETRO","") or "").strip()
            return cd if cd else "CONTROL"

        ctrl_consol["CÓDIGO DE USUARIO"] = ctrl_consol.apply(_fill_usercode, axis=1)
        ctrl_consol["NOMBRE"] = "CONTROL"

    out = pd.concat([ctrl_consol, per_consol], ignore_index=True, sort=False)
    if out.empty:
        return out

    for h in ["Hp (10)","Hp (0.07)","Hp (3)"]:
        out[h] = out[h].map(lambda v: pmfmt2(v, umbral_pm))

    orden_pref = [
        "PERIODO DE LECTURA","CLIENTE","CÓDIGO DE DOSÍMETRO","CÓDIGO DE USUARIO","NOMBRE",
        "CÉDULA","FECHA DE LECTURA","TIPO DE DOSÍMETRO","Hp (10)","Hp (0.07)","Hp (3)"
    ]
    cols = [c for c in orden_pref if c in out.columns] + [c for c in out.columns if c not in orden_pref]
    out = out[cols]
    sort_keys = [c for c in ["PERIODO DE LECTURA","NOMBRE","CÓDIGO DE USUARIO","CÓDIGO DE DOSÍMETRO"] if c in out.columns]
    out = out.sort_values(sort_keys).reset_index(drop=True)
    return out

# ===================== Excel builder =====================
def draw_box(ws, r1, c1, r2, c2, color="000000", style="thin"):
    """Dibuja un recuadro alrededor de [r1:c1]..[r2:c2] sin tocar bordes internos."""
    side = Side(style=style, color=color)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left   = side if c == c1 else cell.border.left,
                right  = side if c == c2 else cell.border.right,
                top    = side if r == r1 else cell.border.top,
                bottom = side if r == r2 else cell.border.bottom,
            )

def build_excel_like_example(
    df_reporte: pd.DataFrame,
    fecha_emision: str,
    cliente: str,
    codigo_reporte: str,
    logo_bytes: Optional[bytes] = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "REPORTE"

    # Anchos de columna (A..N aprox)
    widths = [22,18,18,18,18,18,18,18,18,18,18,18,18,18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Estilos básicos
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold   = Font(bold=True, size=12)
    big    = Font(bold=True, size=14)

    # Encabezado en A1 (multilínea)
    header_text = "MICROSIEVERT, S.A.\nPH Conardo\nCalle 41 Este, Panamá\nPANAMÁ"
    ws["A1"].value = header_text
    ws["A1"].font = big
    ws["A1"].alignment = left
    ws.row_dimensions[1].height = 70  # para acomodar líneas

    # Logo en E1 (puedes cambiar a F1 si prefieres)
    if logo_bytes:
        img = XLImage(BytesIO(logo_bytes))
        img.width = 180
        img.height = 60
        img.anchor = "E1"   # o "F1"
        ws.add_image(img)

    # Cuadro Fecha/Cliente/Código en K2:N4
    start_r, end_r = 2, 4
    start_c, end_c = 11, 14  # K..N
    labels = [("Fecha de emisión", fecha_emision), ("Cliente", cliente), ("Código", codigo_reporte)]
    rr = start_r
    for label, value in labels:
        ws.cell(row=rr, column=start_c, value=label).alignment = left
        ws.cell(row=rr, column=start_c).font = bold
        ws.merge_cells(start_row=rr, start_column=start_c + 1, end_row=rr, end_column=end_c)
        vcell = ws.cell(row=rr, column=start_c + 1, value=value)
        vcell.alignment = left
        rr += 1
    draw_box(ws, start_r, start_c, end_r, end_c, color="000000", style="thin")

    # Título tabla de resultados
    start_table_row = 6
    ws.merge_cells(start_row=start_table_row, start_column=1, end_row=start_table_row, end_column=14)
    ws.cell(row=start_table_row, column=1, value="REPORTE DE DOSIMETRÍA").alignment = center
    ws.cell(row=start_table_row, column=1).font = Font(bold=True, size=13)

    # Cabecera de la tabla (según orden solicitado)
    table_cols = [
        "CÓDIGO DE USUARIO","CLIENTE","NOMBRE","CÉDULA",
        "Hp (10)","Hp (0.07)","Hp (3)",
        "Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL",
        "Hp (10) DE POR VIDA","Hp (0.07) DE POR VIDA","Hp (3) DE POR VIDA",
        "CÓDIGO DE DOSÍMETRO"  # añadido al final según pedido
    ]
    header_row = start_table_row + 2
    for j, colname in enumerate(table_cols, start=1):
        cell = ws.cell(row=header_row, column=j, value=colname)
        cell.alignment = center
        cell.font = bold
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Filas de datos
    data_row = header_row + 1
    if not df_reporte.empty:
        # Asegurar columnas presentes
        for c in table_cols:
            if c not in df_reporte.columns:
                df_reporte[c] = ""
        # Escribir
        for _, row in df_reporte[table_cols].iterrows():
            for j, colname in enumerate(table_cols, start=1):
                val = row[colname]
                ws.cell(row=data_row, column=j, value=val)
            data_row += 1

    # Marco de la tabla completa (cabecera + datos)
    if data_row > header_row + 1:
        draw_box(ws, header_row, 1, data_row - 1, len(table_cols), style="thin")

    # ===== Bloque de información debajo de la tabla =====
    info_start = data_row + 1
    # Título bloque
    ws.merge_cells(start_row=info_start, start_column=1, end_row=info_start, end_column=14)
    ws.cell(row=info_start, column=1, value="INFORMACIÓN DEL REPORTE DE DOSIMETRÍA").font = Font(bold=True, size=12)
    ws.cell(row=info_start, column=1).alignment = left

    # Sub-bloque: Periodo / Fecha lectura / Tipo dosímetro
    blk1_r1 = info_start + 2
    ws.cell(row=blk1_r1, column=1, value="– Periodo de lectura:").font = bold
    ws.merge_cells(start_row=blk1_r1, start_column=2, end_row=blk1_r1, end_column=14)
    ws.cell(row=blk1_r1, column=2, value="periodo de uso del dosímetro personal.").alignment = left

    blk1_r2 = blk1_r1 + 1
    ws.cell(row=blk1_r2, column=1, value="– Fecha de lectura:").font = bold
    ws.merge_cells(start_row=blk1_r2, start_column=2, end_row=blk1_r2, end_column=14)
    ws.cell(row=blk1_r2, column=2, value="corresponde a la fecha en que fue realizada la lectura del dosímetro.").alignment = left

    blk1_r3 = blk1_r2 + 1
    ws.cell(row=blk1_r3, column=1, value="– Tipo de dosímetro:").font = bold
    ws.merge_cells(start_row=blk1_r3, start_column=2, end_row=blk1_r3, end_column=14)
    ws.cell(row=blk1_r3, column=2, value="CE = Cuerpo Entero, A = Anillo, B = Brazalete, CR = Cristalino").alignment = left

    # Enmarcar sub-bloque 1
    draw_box(ws, blk1_r1, 1, blk1_r3, 14, style="thin")

    # Sub-bloque: Datos del participante
    blk2_r1 = blk1_r3 + 2
    ws.cell(row=blk2_r1, column=1, value="– DATOS DEL PARTICIPANTE:").font = bold
    ws.merge_cells(start_row=blk2_r1, start_column=1, end_row=blk2_r1, end_column=14)
    ws.cell(row=blk2_r1, column=1).alignment = left

    blk2_r2 = blk2_r1 + 1
    ws.merge_cells(start_row=blk2_r2, start_column=1, end_row=blk2_r2, end_column=14)
    ws.cell(row=blk2_r2, column=1, value="- Código de usuario: Número único asignado al usuario por Microsievert, S.A.").alignment = left

    blk2_r3 = blk2_r2 + 1
    ws.merge_cells(start_row=blk2_r3, start_column=1, end_row=blk2_r3, end_column=14)
    ws.cell(row=blk2_r3, column=1, value="- Nombre: Persona a la cual se le asigna el dosímetro personal.").alignment = left

    blk2_r4 = blk2_r3 + 1
    ws.merge_cells(start_row=blk2_r4, start_column=1, end_row=blk2_r4, end_column=14)
    ws.cell(row=blk2_r4, column=1, value="- Cédula: Número del documento de identidad personal del usuario.").alignment = left

    blk2_r5 = blk2_r4 + 1
    ws.merge_cells(start_row=blk2_r5, start_column=1, end_row=blk2_r5, end_column=14)
    ws.cell(row=blk2_r5, column=1, value="- Fecha de nacimiento: Registro de la fecha de nacimiento del usuario.").alignment = left

    draw_box(ws, blk2_r1, 1, blk2_r5, 14, style="thin")

    # Sub-bloque: Dosis en mSv (tabla pequeña)
    blk3_r1 = blk2_r5 + 2
    ws.cell(row=blk3_r1, column=1, value="– DOSIS EN MILISIEVERT:").font = bold
    ws.merge_cells(start_row=blk3_r1, start_column=1, end_row=blk3_r1, end_column=14)
    ws.cell(row=blk3_r1, column=1).alignment = left

    # Cabeceras mini tabla
    mini_header = ["Nombre","Definición","Unidad"]
    mini_rh = blk3_r1 + 1
    for j, t in enumerate(mini_header, start=1):
        cell = ws.cell(row=mini_rh, column=j, value=t)
        cell.font = bold
        cell.alignment = center
        cell.fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")

    # Filas mini tabla
    rows_info = [
        ("Dosis efectiva", "Es la dosis equivalente en tejido blando a 10 mm de profundidad.", "mSv"),
        ("Dosis equivalente superficial", "Dosis equivalente a 0,07 mm de profundidad en tejido blando.", "mSv"),
        ("Dosis equivalente a cristalino", "Dosis equivalente a 3 mm de profundidad.", "mSv"),
    ]
    rr = mini_rh + 1
    for (n, d, u) in rows_info:
        ws.cell(row=rr, column=1, value=n).alignment = left
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=13)
        ws.cell(row=rr, column=2, value=d).alignment = left
        ws.cell(row=rr, column=14, value=u).alignment = center
        rr += 1
    draw_box(ws, mini_rh, 1, rr-1, 14, style="thin")

    # Sub-bloque: Aclaraciones finales
    blk4_r1 = rr + 2
    ws.merge_cells(start_row=blk4_r1, start_column=1, end_row=blk4_r1, end_column=14)
    ws.cell(row=blk4_r1, column=1, value="LECTURAS DE ANILLO: registradas como dosis equivalente superficial Hp(0.07).").alignment = left

    blk4_r2 = blk4_r1 + 1
    ws.merge_cells(start_row=blk4_r2, start_column=1, end_row=blk4_r2, end_column=14)
    ws.cell(row=blk4_r2, column=1, value="DOSÍMETRO DE CONTROL: incluido en cada paquete para monitorear exposición durante tránsito y almacenamiento.").alignment = left

    blk4_r3 = blk4_r2 + 1
    ws.merge_cells(start_row=blk4_r3, start_column=1, end_row=blk4_r3, end_column=14)
    ws.cell(row=blk4_r3, column=1, value="POR DEBAJO DEL MÍNIMO DETECTADO (PM): dosis por debajo del mínimo reportable para el periodo.").alignment = left

    draw_box(ws, blk4_r1, 1, blk4_r3, 14, style="thin")

    # Guardar
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ===================== REPORTE ÚNICO (CONTROL primero) =====================
def construir_reporte_unico(
    df_vista: pd.DataFrame,
    df_num: pd.DataFrame,
    umbral_pm: float = 0.005,
    agrupar_control_por: str = "CLIENTE",
) -> pd.DataFrame:
    if df_vista is None or df_vista.empty or df_num is None or df_num.empty:
        return pd.DataFrame()

    # PERSONAS
    personas_num = df_num[~df_num["NOMBRE"].apply(is_control_name)].copy()
    if not personas_num.empty:
        per_anual = personas_num.groupby("CÓDIGO DE USUARIO", as_index=False).agg({
            "CLIENTE":"last","NOMBRE":"last","CÉDULA":"last","CÓDIGO DE DOSÍMETRO":"last",
            "_Hp10_NUM":"sum","_Hp007_NUM":"sum","_Hp3_NUM":"sum"
        }).rename(columns={"_Hp10_NUM":"Hp (10) ANUAL","_Hp007_NUM":"Hp (0.07) ANUAL","_Hp3_NUM":"Hp (3) ANUAL"})
        personas_num["__fecha__"] = personas_num["PERIODO DE LECTURA"].map(periodo_to_date)
        idx_last = personas_num.groupby("CÓDIGO DE USUARIO")["__fecha__"].idxmax()
        per_last = personas_num.loc[idx_last, safe_cols(personas_num, [
            "CÓDIGO DE USUARIO","PERIODO DE LECTURA","_Hp10_NUM","_Hp007_NUM","_Hp3_NUM",
            "FECHA DE LECTURA","TIPO DE DOSÍMETRO"
        ])].rename(columns={"_Hp10_NUM":"Hp (10)","_Hp007_NUM":"Hp (0.07)","_Hp3_NUM":"Hp (3)"})
        per_view = per_anual.merge(per_last, on="CÓDIGO DE USUARIO", how="left")
        for c in safe_cols(per_view, ["Hp (10)","Hp (0.07)","Hp (3)","Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL"]):
            per_view[c] = per_view[c].map(lambda v: pmfmt2(v, umbral_pm))
        per_view["Hp (10) DE POR VIDA"]   = per_view["Hp (10) ANUAL"]
        per_view["Hp (0.07) DE POR VIDA"] = per_view["Hp (0.07) ANUAL"]
        per_view["Hp (3) DE POR VIDA"]    = per_view["Hp (3) ANUAL"]
        personas_final = per_view[safe_cols(per_view, [
            "PERIODO DE LECTURA","CLIENTE","CÓDIGO DE DOSÍMETRO","CÓDIGO DE USUARIO","NOMBRE","CÉDULA",
            "FECHA DE LECTURA","TIPO DE DOSÍMETRO",
            "Hp (10)","Hp (0.07)","Hp (3)",
            "Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL",
            "Hp (10) DE POR VIDA","Hp (0.07) DE POR VIDA","Hp (3) DE POR VIDA"
        ])]
    else:
        personas_final = pd.DataFrame(columns=[
            "PERIODO DE LECTURA","CLIENTE","CÓDIGO DE DOSÍMETRO","CÓDIGO DE USUARIO","NOMBRE","CÉDULA",
            "FECHA DE LECTURA","TIPO DE DOSÍMETRO",
            "Hp (10)","Hp (0.07)","Hp (3)",
            "Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL",
            "Hp (10) DE POR VIDA","Hp (0.07) DE POR VIDA","Hp (3) DE POR VIDA"
        ])

    # CONTROL (una fila por CLIENTE)
    control_v = df_vista[df_vista["NOMBRE"].apply(is_control_name)].copy()
    if not control_v.empty:
        for h in safe_cols(control_v, ["Hp (10)","Hp (0.07)","Hp (3)"]):
            control_v[h] = control_v[h].apply(hp_to_num)

        agr = agrupar_control_por if agrupar_control_por in control_v.columns else None
        if agr is None:
            control_v["__grupo__"] = "GLOBAL"
            agr = "__grupo__"

        ctrl_anual = control_v.groupby(agr, as_index=False).agg({
            "CLIENTE":"last",
            "Hp (10)":"sum","Hp (0.07)":"sum","Hp (3)":"sum"
        }).rename(columns={"Hp (10)":"Hp (10) ANUAL","Hp (0.07)":"Hp (0.07) ANUAL","Hp (3)":"Hp (3) ANUAL"})

        tmp = control_v.copy()
        tmp["__fecha__"] = tmp["PERIODO DE LECTURA"].map(periodo_to_date)
        idx_last_c = tmp.groupby(agr)["__fecha__"].idxmax()
        last_vals = tmp.loc[idx_last_c, safe_cols(tmp, [
            agr,"PERIODO DE LECTURA","Hp (10)","Hp (0.07)","Hp (3)",
            "CÓDIGO DE DOSÍMETRO","CÓDIGO DE USUARIO","CÉDULA","FECHA DE LECTURA","TIPO DE DOSÍMETRO"
        ])]

        ctrl_view = ctrl_anual.merge(last_vals, on=agr, how="left")
        ctrl_view["NOMBRE"] = "CONTROL"

        def _fill_usercode(row):
            cu = str(row.get("CÓDIGO DE USUARIO","") or "").strip()
            return cu if cu else str(row.get("CÓDIGO DE DOSÍMETRO","") or "").strip()
        ctrl_view["CÓDIGO DE USUARIO"] = ctrl_view.apply(_fill_usercode, axis=1)

        for c in safe_cols(ctrl_view, ["Hp (10)","Hp (0.07)","Hp (3)","Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL"]):
            ctrl_view[c] = ctrl_view[c].map(lambda v: pmfmt2(v, umbral_pm))
        ctrl_view["Hp (10) DE POR VIDA"]   = ctrl_view["Hp (10) ANUAL"]
        ctrl_view["Hp (0.07) DE POR VIDA"] = ctrl_view["Hp (0.07) ANUAL"]
        ctrl_view["Hp (3) DE POR VIDA"]    = ctrl_view["Hp (3) ANUAL"]

        ctrl_final = ctrl_view[safe_cols(ctrl_view, [
            "PERIODO DE LECTURA","CLIENTE","CÓDIGO DE DOSÍMETRO","CÓDIGO DE USUARIO","NOMBRE","CÉDULA",
            "FECHA DE LECTURA","TIPO DE DOSÍMETRO",
            "Hp (10)","Hp (0.07)","Hp (3)",
            "Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL",
            "Hp (10) DE POR VIDA","Hp (0.07) DE POR VIDA","Hp (3) DE POR VIDA"
        ])]
    else:
        ctrl_final = pd.DataFrame(columns=personas_final.columns)

    # Unión CONTROL primero
    reporte = pd.concat([ctrl_final, personas_final], ignore_index=True)
    if not reporte.empty:
        reporte["__is_control__"] = reporte["NOMBRE"].apply(is_control_name)
        # Orden por CODIGO DE DOSÍMETRO si quieres ordenarlo así (puedes ajustar):
        reporte = reporte.sort_values(
            by=["__is_control__","CÓDIGO DE DOSÍMETRO","CÓDIGO DE USUARIO","NOMBRE"],
            ascending=[False, True, True, True]
        ).drop(columns=["__is_control__"])
    return reporte

# ===================== UI: Tabs =====================
tab1, tab2 = st.tabs(["1) Cargar y Subir a Ninox", "2) Reporte Final (sumas)"])

# ------------------ TAB 1 ------------------
with tab1:
    st.subheader("1) Cargar LISTA DE CÓDIGO")
    upl_lista = st.file_uploader("Sube la LISTA DE CÓDIGO (CSV / XLS / XLSX)", type=["csv","xls","xlsx"], key="upl_lista")
    df_lista = leer_lista_codigo(upl_lista) if upl_lista else None
    if df_lista is not None and not df_lista.empty:
        st.success(f"LISTA cargada: {len(df_lista)} filas")
        st.dataframe(df_lista.head(20), use_container_width=True)
    else:
        st.info("LISTA vacía o sin datos")

    st.subheader("2) Subir Archivo de Dosis")
    upl_dosis = st.file_uploader("Selecciona CSV/XLS/XLSX (dosis)", type=["csv","xls","xlsx"], key="upl_dosis")
    df_dosis = leer_dosis(upl_dosis) if upl_dosis else None
    if df_dosis is not None and not df_dosis.empty:
        st.success(f"Dosis cargadas: {len(df_dosis)} fila(s)")
        st.dataframe(df_dosis.head(15), use_container_width=True)

    per_options = sorted(df_lista["PERIODO DE LECTURA"].dropna().astype(str).str.upper().unique().tolist()) if df_lista is not None else []
    periodos_sel = st.multiselect("Filtrar por PERIODO DE LECTURA (elige uno o varios; vacío = TODOS)", per_options, default=[])

    with st.expander("⚙️ Opcional: Control manual si NO existe CONTROL en el periodo"):
        use_manual_ctrl = st.checkbox("Activar control manual", value=False)
        manual_ctrl_val = st.number_input("Valor de control manual a restar (Hp10, Hp0.07, Hp3)", min_value=0.0, step=0.001, format="%.3f", value=0.000)

    subir_pm_como_texto = st.checkbox("Guardar 'PM' como texto en Ninox (si desmarcas, sube None en PM)", value=True)

    if st.button("✅ Procesar y Previsualizar", type="primary"):
        if df_lista is None or df_lista.empty:
            st.error("Primero sube la LISTA DE CÓDIGO.")
        elif df_dosis is None or df_dosis.empty:
            st.error("Sube el archivo de dosis.")
        elif "dosimeter" not in df_dosis.columns:
            st.error("El archivo de dosis debe incluir la columna 'dosimeter'.")
        else:
            df_final_raw = construir_registros(df_lista, df_dosis, periodos_sel)
            if df_final_raw.empty:
                with st.expander("Debug de coincidencias (no se encontraron)"):
                    st.write({
                        "dosimeter únicos en dosis": sorted(df_dosis["dosimeter"].dropna().unique().tolist()) if "dosimeter" in df_dosis.columns else [],
                        "CÓDIGO_DOSÍMETRO únicos en LISTA (según filtro)": sorted(df_lista["CÓDIGO_DOSÍMETRO"].dropna().unique().tolist()) if "CÓDIGO_DOSÍMETRO" in df_lista.columns else []
                    })
                st.warning("⚠️ No hay coincidencias **CÓDIGO_DOSÍMETRO** ↔ **dosimeter** (revisa periodos/códigos).")
            else:
                df_vista, df_num_corr = aplicar_resta_control_y_formato(
                    df_final_raw, umbral_pm=0.005,
                    manual_ctrl=(manual_ctrl_val if use_manual_ctrl else None)
                )
                st.session_state.df_final_vista = df_vista.drop(columns=["_IS_CONTROL"], errors="ignore")
                st.session_state.df_final_num   = df_num_corr
                st.success(f"¡Listo! Registros generados (corregidos): {len(st.session_state.df_final_vista)}")
                st.dataframe(st.session_state.df_final_vista, use_container_width=True)
                st.caption(f"Controles detectados: {(st.session_state.df_final_vista['NOMBRE'].apply(is_control_name)).sum()}")

    st.markdown("---")
    st.subheader("3) Subir TODO a Ninox (tabla **BASE DE DATOS**)")

    def _to_str(v):
        if pd.isna(v): return ""
        if isinstance(v, (pd.Timestamp, )):
            return v.strftime("%Y-%m-%d %H:%M:%S")
        return str(v)

    def _hp_value_for_upload(v, as_text_pm=True):
        if isinstance(v, str) and v.strip().upper() == "PM":
            return "PM" if as_text_pm else None
        try:
            num = float(v)
        except Exception:
            return v if v is not None else None
        return f"{num:.2f}" if as_text_pm else num

    if st.button("⬆️ Subir a Ninox (BASE DE DATOS)"):
        df_vista = st.session_state.get("df_final_vista")
        df_num   = st.session_state.get("df_final_num")
        if df_vista is None or df_vista.empty or df_num is None or df_num.empty:
            st.error("No hay datos procesados. Pulsa 'Procesar y Previsualizar' primero.")
        else:
            df_para_subir = consolidar_para_upload(df_vista, df_num, umbral_pm=0.005)
            if df_para_subir.empty:
                st.error("Nada para subir después de consolidar.")
            else:
                rows = []
                for _, row in df_para_subir.iterrows():
                    fields = {
                        "PERIODO DE LECTURA": _to_str(row.get("PERIODO DE LECTURA","")),
                        "CLIENTE": _to_str(row.get("CLIENTE","")),
                        "CÓDIGO DE DOSÍMETRO": _to_str(row.get("CÓDIGO DE DOSÍMETRO","")),
                        "CÓDIGO DE USUARIO": _to_str(row.get("CÓDIGO DE USUARIO","")),
                        "NOMBRE": _to_str(row.get("NOMBRE","")),
                        "CÉDULA": _to_str(row.get("CÉDULA","")),
                        "FECHA DE LECTURA": _to_str(row.get("FECHA DE LECTURA","")),
                        "TIPO DE DOSÍMETRO": _to_str(row.get("TIPO DE DOSÍMETRO","") or "CE"),
                        "Hp (10)": _hp_value_for_upload(row.get("Hp (10)"), subir_pm_como_texto),
                        "Hp (0.07)": _hp_value_for_upload(row.get("Hp (0.07)"), subir_pm_como_texto),
                        "Hp (3)": _hp_value_for_upload(row.get("Hp (3)"), subir_pm_como_texto),
                    }
                    rows.append({"fields": fields})
                with st.spinner("Subiendo a Ninox..."):
                    res = ninox_insert(TABLE_WRITE_NAME, rows, batch_size=300)
                if res.get("ok"):
                    st.success(f"✅ Subido a Ninox: {res.get('inserted', 0)} registro(s).")
                    st.toast("¡Datos enviados a Ninox!", icon="✅")
                else:
                    st.error(f"❌ Error al subir: {res.get('error')}")

# ------------------ TAB 2 ------------------
with tab2:
    st.subheader("📊 Reporte Final (CONTROL primero y luego PERSONAS)")

    fuente = st.radio("Fuente de datos para el reporte:", [
        "Usar datos procesados en esta sesión",
        "Leer directamente de Ninox (tabla BASE DE DATOS)",
    ], index=0)

    # Subida de logo
    upl_logo = st.file_uploader("Logo (opcional, PNG/JPG)", type=["png","jpg","jpeg"], key="upl_logo")

    if fuente == "Leer directamente de Ninox (tabla BASE DE DATOS)":
        try:
            with st.spinner("Leyendo registros desde Ninox…"):
                recs = ninox_list_records(TABLE_WRITE_NAME, limit=1000)
                df_nx = ninox_records_to_df(recs)
            if df_nx.empty:
                st.warning("No se recibieron registros desde Ninox.")
            else:
                st.session_state.df_final_vista = df_nx.copy()
                tmp = df_nx.copy()
                for h in ["Hp (10)","Hp (0.07)","Hp (3)"]:
                    tmp[h] = tmp[h].apply(hp_to_num)
                tmp["_Hp10_NUM"]  = tmp["Hp (10)"]
                tmp["_Hp007_NUM"] = tmp["Hp (0.07)"]
                tmp["_Hp3_NUM"]   = tmp["Hp (3)"]
                st.session_state.df_final_num = tmp[[
                    "_Hp10_NUM","_Hp007_NUM","_Hp3_NUM","PERIODO DE LECTURA","CLIENTE",
                    "CÓDIGO DE USUARIO","CÓDIGO DE DOSÍMETRO","NOMBRE","CÉDULA",
                    "TIPO DE DOSÍMETRO","FECHA DE LECTURA"
                ]].copy()
        except Exception as e:
            st.error(f"Error leyendo Ninox: {e}")

    df_vista = st.session_state.get("df_final_vista")
    df_num   = st.session_state.get("df_final_num")
    if df_vista is None or df_vista.empty or df_num is None or df_num.empty:
        st.info("No hay datos para mostrar en el reporte final.")
    else:
        # Filtro por Cliente
        clientes = sorted([c for c in df_vista["CLIENTE"].dropna().unique().tolist() if str(c).strip()])
        cliente_sel = st.selectbox("Filtrar por CLIENTE (opcional)", ["(Todos)"] + clientes, index=0)
        if cliente_sel != "(Todos)":
            df_vista = df_vista[df_vista["CLIENTE"] == cliente_sel].copy()
            df_num   = df_num[df_num["CLIENTE"] == cliente_sel].copy()

        # Construimos la tabla única
        reporte = construir_reporte_unico(df_vista, df_num, umbral_pm=0.005, agrupar_control_por="CLIENTE")
        if reporte.empty:
            st.info("No hay datos para el reporte con el filtro aplicado.")
        else:
            st.dataframe(reporte, use_container_width=True)

            # Descarga CSV
            csv_bytes = reporte.to_csv(index=False).encode("utf-8-sig")
            st.download_button("⬇️ Descargar CSV", data=csv_bytes, file_name="Reporte_Final.csv", mime="text/csv")

            # Descarga Excel con formato tipo ejemplo
            fecha_emision = datetime.now().strftime("%d/%m/%Y")
            cliente_val   = cliente_sel if cliente_sel != "(Todos)" else (reporte["CLIENTE"].iloc[0] if "CLIENTE" in reporte.columns and not reporte["CLIENTE"].isna().all() else "")
            codigo_ui     = st.text_input("Código del reporte (opcional)", value="SIN-CÓDIGO")

            logo_bytes = upl_logo.read() if upl_logo is not None else None
            excel_bytes = build_excel_like_example(
                df_reporte=reporte,
                fecha_emision=fecha_emision,
                cliente=cliente_val,
                codigo_reporte=codigo_ui or "SIN-CÓDIGO",
                logo_bytes=logo_bytes,
            )
            st.download_button(
                "⬇️ Descargar Excel (formato ejemplo)",
                data=excel_bytes,
                file_name="Reporte_Final_Formato.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

